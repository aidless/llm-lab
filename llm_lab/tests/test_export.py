"""Tests for export.py — JSON, CSV, XLSX, and HTML exporters."""

import json
import re

from llm_lab import export

_SAMPLE_ROWS = [
    {"id": 1, "intent_id": "abc", "seq": 1, "timestamp": "2025-01-01", "action": "call", "model": "gpt-4o", "detail": "hello", "cost_usd": 0.001},
    {"id": 2, "intent_id": "abc", "seq": 2, "timestamp": "2025-01-01", "action": "verdict", "model": None, "detail": None, "cost_usd": None},
]


class TestExportJson:
    def test_returns_valid_json(self):
        raw = export.export_json("abc", _SAMPLE_ROWS)
        data = json.loads(raw)
        assert data["intent_id"] == "abc"
        assert len(data["events"]) == 2

    def test_events_serialized(self):
        raw = export.export_json("abc", _SAMPLE_ROWS)
        data = json.loads(raw)
        assert data["events"][0]["action"] == "call"

    def test_empty_events(self):
        raw = export.export_json("empty", [])
        data = json.loads(raw)
        assert data["intent_id"] == "empty"
        assert data["events"] == []

    def test_non_ascii_handling(self):
        rows = [{"id": 1, "intent_id": "cn", "seq": 1, "timestamp": "2025-01-01", "action": "测试", "model": None, "detail": "中文", "cost_usd": None}]
        raw = export.export_json("cn", rows)
        assert "测试" in raw
        assert "\\u" not in raw


class TestExportCsv:
    def test_header_present(self):
        raw = export.export_csv(_SAMPLE_ROWS)
        lines = raw.strip().splitlines()
        assert lines[0] == "id,intent_id,seq,timestamp,action,model,detail,cost_usd"

    def test_data_rows(self):
        raw = export.export_csv(_SAMPLE_ROWS)
        lines = raw.strip().splitlines()
        assert len(lines) == 3
        assert lines[1] == "1,abc,1,2025-01-01,call,gpt-4o,hello,0.001"

    def test_null_values_as_empty(self):
        raw = export.export_csv(_SAMPLE_ROWS)
        lines = raw.strip().splitlines()
        assert re.search(r",,,$", lines[2])

    def test_empty_rows(self):
        raw = export.export_csv([])
        assert raw.strip() == "id,intent_id,seq,timestamp,action,model,detail,cost_usd"


class TestExportXlsx:
    def test_requires_openpyxl_direct(self):
        """Trigger the ImportError path by making openpyxl unimportable."""
        import builtins
        real_import = builtins.__import__

        def _mock_import(name, *args, **kw):
            if name == "openpyxl":
                raise ImportError("no openpyxl")
            return real_import(name, *args, **kw)

        builtins.__import__ = _mock_import
        try:
            with __import__("pytest").raises(RuntimeError, match="openpyxl is required"):
                export.export_xlsx(_SAMPLE_ROWS)
        finally:
            builtins.__import__ = real_import

    def test_requires_openpyxl_skip(self):
        """openpyxl IS installed so we skip (ImportError path tested by test_requires_openpyxl_direct)."""
        __import__("pytest").importorskip("openpyxl")
        __import__("pytest").skip("openpyxl is installed — tested by test_requires_openpyxl_direct")

    def test_returns_xlsx_bytes(self):
        __import__("pytest").importorskip("openpyxl")
        content = export.export_xlsx(_SAMPLE_ROWS)
        assert isinstance(content, bytes)
        assert len(content) > 0
        assert content[:2] == b"PK"

    def test_xlsx_contains_header_and_data(self):
        __import__("pytest").importorskip("openpyxl")
        import io  # noqa: I001
        import openpyxl
        content = export.export_xlsx(_SAMPLE_ROWS)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        assert ws.title == "Events"
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("id", "intent_id", "seq", "timestamp", "action", "model", "detail", "cost_usd")
        assert rows[1] == (1, "abc", 1, "2025-01-01", "call", "gpt-4o", "hello", 0.001)

    def test_xlsx_empty_rows(self):
        __import__("pytest").importorskip("openpyxl")
        import io  # noqa: I001
        import openpyxl
        content = export.export_xlsx([])
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1
        assert rows[0] == ("id", "intent_id", "seq", "timestamp", "action", "model", "detail", "cost_usd")


class TestExportHtml:
    def test_returns_html_string(self):
        result = {
            "intent_id": "test-123", "goal": "test goal", "model": "gpt-4o",
            "steps": 1, "total_tokens": 10, "total_cost_usd": 0.001, "all_passed": True,
            "steps_detail": [
                {"action": "test", "verdict": {"label": "pass"}, "tokens": 10, "cost": 0.001},
            ],
        }
        html = export.export_html(result)
        assert html.startswith("<!DOCTYPE html>")
        assert "test-123" in html

    def test_contains_intent_id_and_model(self):
        result = {
            "intent_id": "abc-xyz", "goal": "translate", "model": "claude-3",
            "steps": 2, "total_tokens": 20, "total_cost_usd": 0.002, "all_passed": True,
            "steps_detail": [],
        }
        html = export.export_html(result)
        assert "abc-xyz" in html
        assert "claude-3" in html
        assert "translate" in html

    def test_fail_status(self):
        result = {
            "intent_id": "fail-run", "goal": "test", "model": "gpt-4o",
            "steps": 1, "total_tokens": 5, "total_cost_usd": 0.0005, "all_passed": False,
            "steps_detail": [
                {"action": "test", "verdict": {"label": "fail"}, "tokens": 5, "cost": 0.0005},
            ],
        }
        html = export.export_html(result)
        assert "FAIL" in html

    def test_empty_steps(self):
        result = {
            "intent_id": "empty", "goal": "", "model": "",
            "steps": 0, "total_tokens": 0, "total_cost_usd": 0.0, "all_passed": False,
            "steps_detail": [],
        }
        html = export.export_html(result)
        assert "</table>" in html
        assert "FAIL" in html

    def test_displays_verdict_badges(self):
        result = {
            "intent_id": "badges", "goal": "", "model": "gpt-4o",
            "steps": 3, "total_tokens": 30, "total_cost_usd": 0.003, "all_passed": False,
            "steps_detail": [
                {"action": "a", "verdict": {"label": "pass"}, "tokens": 10, "cost": 0.001},
                {"action": "b", "verdict": {"label": "fail"}, "tokens": 10, "cost": 0.001},
                {"action": "c", "verdict": {"label": "partial"}, "tokens": 10, "cost": 0.001},
            ],
        }
        html = export.export_html(result)
        assert "PASS" in html
        assert "FAIL" in html
        assert "PARTIAL" in html

    def test_json_detail_block_present(self):
        result = {
            "intent_id": "json-block", "goal": "", "model": "",
            "steps": 0, "total_tokens": 0, "total_cost_usd": 0.0, "all_passed": True,
            "steps_detail": [],
        }
        html = export.export_html(result)
        assert "<pre>" in html

    def test_metric_check_in_steps_detail(self):
        result = {
            "intent_id": "metrics", "goal": "", "model": "gpt-4o",
            "steps": 1, "total_tokens": 10, "total_cost_usd": 0.001, "all_passed": False,
            "steps_detail": [
                {
                    "action": "test", "verdict": {"label": "fail"},
                    "tokens": 10, "cost": 0.001,
                    "metric_check": {"label": "fail", "reason": "output too short"},
                },
            ],
        }
        html = export.export_html(result)
        assert "metric_check" in html
        assert "output too short" in html
